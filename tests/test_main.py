import os
import pandas as pd
import subprocess
import openpyxl
import shlex
import numpy as np

# Define file paths
INPUT_FILE = "test_input.xlsx"
INPUT_FILE_2 = "test_input2.xlsx"
OUTPUT_FILE = "test_output.xlsx"

def setup_module(module):
    """Set up dummy Excel files for testing."""
    # Main input file for general tests and data_validation
    data = {'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Eva', 'Frank'],
            'Department': ['Sales', 'Marketing', 'Engineering', 'Sales', 'Engineering', np.nan],
            'Salary': [70000, 65000, 90000, 72000, 95000, 80000],
            'Rating': [5, 4, 5, 3, 4, np.nan]}
    df = pd.DataFrame(data, dtype=object) # Create DataFrame with object dtype to prevent premature type inference
    df.to_excel(INPUT_FILE, index=False, sheet_name='Employees')

    # Second input file for merge action
    data2 = {'Department': ['Sales', 'Marketing', 'Engineering'],
             'Location': ['New York', 'London', 'Paris']}
    df2 = pd.DataFrame(data2)
    df2.to_excel(INPUT_FILE_2, index=False)

def teardown_module(module):
    """Clean up dummy Excel files after testing."""
    files_to_remove = [INPUT_FILE, INPUT_FILE_2, OUTPUT_FILE, "merged_output.xlsx", "sorted_output.xlsx",
                       "renamed_output.xlsx", "deduplicated_output.xlsx", "duplicated_sheet_output.xlsx",
                       "updated_cells_output.xlsx", "summarized_output.xlsx", "calculated_output.xlsx",
                       "sample_data.xlsx", "sales_report.xlsx", "department_salaries.xlsx", "bonus_data.xlsx",
                       "sample_data_with_copy.xlsx", "final_output.xlsx",
                       "filled_na_output.xlsx", "converted_type_output.xlsx"]
    for f in files_to_remove:
        if os.path.exists(f):
            os.remove(f)

def run_cli_command(action, input_file, output_file, sub_action=None, **kwargs):
    """Helper function to run the main.py script."""
    command = ["python3", "main.py", action]
    
    # Place -i and -o arguments right after the main action, before any sub_action
    if action == 'merge':
        command.extend(["--input1", shlex.quote(input_file), "--input2", shlex.quote(kwargs.pop('input2')), "-o", shlex.quote(output_file)])
    else:
        command.extend(["-i", shlex.quote(input_file), "-o", shlex.quote(output_file)])

    if sub_action:
        command.append(sub_action)

    for key, value in kwargs.items():
        # Special handling for list arguments like --by or --subset
        if isinstance(value, list):
            for item in value:
                command.extend([f"--{key.replace("_", "-")}", shlex.quote(str(item))])
        else:
            command.extend([f"--{key.replace("_", "-")}", shlex.quote(str(value))])
    
    # Run the command within the virtual environment
    full_command = f"source venv/bin/activate && {' '.join(command)}"
    result = subprocess.run(["bash", "-c", full_command], capture_output=True, text=True, check=True)
    return result

def test_filter_action():
    """Test the filter action."""
    result = run_cli_command(
        "filter",
        INPUT_FILE,
        OUTPUT_FILE,
        column="Department",
        value="Sales"
    )
    assert result.returncode == 0
    assert "Action 'filter' completed successfully" in result.stdout

    df_output = pd.read_excel(OUTPUT_FILE)
    assert len(df_output) == 2
    assert "Alice" in df_output['Name'].values
    assert "David" in df_output['Name'].values
    assert "Charlie" not in df_output['Name'].values

def test_summarize_action():
    """Test the summarize action."""
    result = run_cli_command(
        "summarize",
        INPUT_FILE,
        "summarized_output.xlsx",
        group_by="Department",
        agg_col="Salary",
        agg_func="mean"
    )
    assert result.returncode == 0
    assert "Action 'summarize' completed successfully" in result.stdout

    df_output = pd.read_excel("summarized_output.xlsx")
    assert len(df_output) == 3
    assert df_output[df_output['Department'] == 'Sales']['Salary'].iloc[0] == 71000.0

def test_calculate_action():
    """Test the calculate action."""
    result = run_cli_command(
        "calculate",
        INPUT_FILE,
        "calculated_output.xlsx",
        new_col="Bonus",
        expr="Salary * 0.1"
    )
    assert result.returncode == 0
    assert "Action 'calculate' completed successfully" in result.stdout

    df_output = pd.read_excel("calculated_output.xlsx")
    assert 'Bonus' in df_output.columns
    assert df_output['Bonus'].iloc[0] == 7000.0

def test_merge_action():
    """Test the merge action."""
    result = run_cli_command(
        "merge",
        INPUT_FILE,
        "merged_output.xlsx",
        input2=INPUT_FILE_2,
        on="Department",
        how="inner"
    )
    assert result.returncode == 0
    assert "Action 'merge' completed successfully" in result.stdout

    df_output = pd.read_excel("merged_output.xlsx")
    assert 'Location' in df_output.columns
    assert len(df_output) == 5

def test_sort_action():
    """Test the sort action."""
    result = run_cli_command(
        "sort",
        INPUT_FILE,
        "sorted_output.xlsx",
        by=["Salary"],
        order="desc"
    )
    assert result.returncode == 0
    assert "Action 'sort' completed successfully" in result.stdout

    df_output = pd.read_excel("sorted_output.xlsx")
    assert df_output['Salary'].iloc[0] == 95000
    assert df_output['Salary'].iloc[-1] == 65000

def test_rename_action():
    """Test the rename action."""
    result = run_cli_command(
        "rename",
        INPUT_FILE,
        "renamed_output.xlsx",
        map="Name:Full Name,Department:Dept"
    )
    assert result.returncode == 0
    assert "Action 'rename' completed successfully" in result.stdout

    df_output = pd.read_excel("renamed_output.xlsx")
    assert 'Full Name' in df_output.columns
    assert 'Dept' in df_output.columns
    assert 'Name' not in df_output.columns
    assert 'Department' not in df_output.columns

def test_drop_duplicates_action():
    """Test the drop_duplicates action."""
    result = run_cli_command(
        "drop_duplicates",
        INPUT_FILE,
        "deduplicated_output.xlsx",
        subset=["Department"]
    )
    assert result.returncode == 0
    assert "Action 'drop_duplicates' completed successfully" in result.stdout

    df_output = pd.read_excel("deduplicated_output.xlsx")
    assert len(df_output) == 4 # Sales, Marketing, Engineering, None

def test_duplicate_sheet_action():
    """Test the duplicate_sheet action."""
    result = run_cli_command(
        "duplicate_sheet",
        INPUT_FILE,
        "duplicated_sheet_output.xlsx",
        source_sheet="Employees",
        new_sheet_name="Employees_Copy"
    )
    assert result.returncode == 0
    assert "Action 'duplicate_sheet' completed successfully" in result.stdout

    workbook = openpyxl.load_workbook("duplicated_sheet_output.xlsx")
    assert "Employees" in workbook.sheetnames
    assert "Employees_Copy" in workbook.sheetnames

def test_update_cells_action():
    """Test the update_cells action."""
    result = run_cli_command(
        "update_cells",
        INPUT_FILE,
        "updated_cells_output.xlsx",
        sheet_name="Employees",
        updates="A1:NewHeaderA,B1:NewHeaderB"
    )
    assert result.returncode == 0
    assert "Action 'update_cells' completed successfully" in result.stdout

    workbook = openpyxl.load_workbook("updated_cells_output.xlsx")
    sheet = workbook['Employees']
    assert sheet['A1'].value == 'NewHeaderA'
    assert sheet['B1'].value == 'NewHeaderB'

def test_data_validation_fill_na_all_columns():
    """Test data_validation fill_na for all columns."""
    result = run_cli_command(
        "data_validation",
        INPUT_FILE,
        "filled_na_output.xlsx",
        sub_action="fill_na",
        value="NIL"
    )
    assert result.returncode == 0
    assert "Action 'data_validation' completed successfully" in result.stdout

    df_output = pd.read_excel("filled_na_output.xlsx", dtype={'Department': str})
    assert df_output['Department'].iloc[5] == 'NIL'

def test_data_validation_fill_na_specific_column():
    """Test data_validation fill_na for a specific column."""
    result = run_cli_command(
        "data_validation",
        INPUT_FILE,
        "filled_na_specific_output.xlsx",
        sub_action="fill_na",
        value="Unknown",
        columns=["Department"]
    )
    assert result.returncode == 0
    assert "Action 'data_validation' completed successfully" in result.stdout

    df_output = pd.read_excel("filled_na_specific_output.xlsx")
    assert df_output['Department'].iloc[5] == 'Unknown'

def test_data_validation_convert_type():
    """Test data_validation convert_type action."""
    result = run_cli_command(
        "data_validation",
        INPUT_FILE,
        "converted_type_output.xlsx",
        sub_action="convert_type",
        column="Rating",
        to_type="int"
    )
    assert result.returncode == 0
    assert "Action 'data_validation' completed successfully" in result.stdout

    df_output = pd.read_excel("converted_type_output.xlsx")
    # After reading from Excel, pandas will use float64 for columns with missing values (NaN)
    # even if they were Int64 before writing. This is expected behavior.
    assert str(df_output['Rating'].dtype) == 'float64'
    assert pd.isna(df_output['Rating'].iloc[5]) # Check if the missing value is still missing
    # Check if non-missing values are correct
    assert df_output['Rating'].iloc[0] == 5.0
    assert df_output['Rating'].iloc[1] == 4.0

def test_chart_action():
    """Test the chart action."""
    result = run_cli_command(
        "chart",
        INPUT_FILE,
        "chart_output.xlsx",
        sheet_name="Employees",
        chart_type="bar",
        x_column="Department",
        y_columns=["Salary"],
        title="Department Salaries",
        chart_title="Salary Chart"
    )
    assert result.returncode == 0
    assert "Action 'chart' completed successfully" in result.stdout

    workbook = openpyxl.load_workbook("chart_output.xlsx")
    assert "Salary Chart" in workbook.sheetnames
    sheet = workbook["Salary Chart"]
    assert sheet._charts
    assert sheet._charts[0].title.tx.rich.p[0].r[0].t == "Department Salaries"
