import argparse
import pandas as pd
import openpyxl

# --- Action Functions ---

def filter_data(df, column, value):
    """Filters the DataFrame based on a column and value."""
    return df[df[column] == value]

def summarize_data(df, group_by_column, agg_column, agg_func):
    """Summarizes the DataFrame by grouping and aggregating."""
    return df.groupby(group_by_column)[agg_column].agg(agg_func).reset_index()

def calculate_column(df, new_column_name, expression):
    """Calculates a new column based on an expression."""
    df[new_column_name] = df.eval(expression)
    return df

def merge_data(df1, df2, on_column, how='inner'):
    """Merges two DataFrames."""
    return pd.merge(df1, df2, on=on_column, how=how)

def sort_data(df, by_columns, ascending=True):
    """Sorts the DataFrame."""
    return df.sort_values(by=by_columns, ascending=ascending)

def rename_columns_data(df, rename_map):
    """Renames columns."""
    return df.rename(columns=rename_map)

def drop_duplicates_data(df, subset=None):
    """Drops duplicate rows."""
    return df.drop_duplicates(subset=subset)

def duplicate_sheet_data(workbook, source_sheet_name, new_sheet_name):
    """Duplicates a sheet in the workbook."""
    source_sheet = workbook[source_sheet_name]
    new_sheet = workbook.copy_worksheet(source_sheet)
    new_sheet.title = new_sheet_name
    return workbook

def update_cells_data(workbook, sheet_name, cell_updates):
    """Updates one or more cells in a specific sheet."""
    sheet = workbook[sheet_name]
    for cell, value in cell_updates.items():
        sheet[cell] = value
    return workbook

def fill_missing_values(df, columns, value):
    """Fills missing values in specified columns."""
    if columns:
        for col in columns:
            df.loc[:, col] = df[col].fillna(value)
    else:
        df = df.fillna(value)
    return df

def convert_column_type(df, column, data_type):
    """Converts the data type of a column."""
    if data_type == 'int':
        # Convert to numeric first, then to nullable integer type
        df[column] = pd.to_numeric(df[column], errors='coerce').astype('Int64')
    elif data_type == 'float':
        df[column] = pd.to_numeric(df[column], errors='coerce')
    elif data_type == 'str':
        df[column] = df[column].astype(str)
    elif data_type == 'datetime':
        df[column] = pd.to_datetime(df[column], errors='coerce')
    else:
        raise ValueError(f"Unsupported data type for conversion: {data_type}")
    return df

def create_chart(workbook, sheet_name, chart_type, x_column, y_columns, title, chart_title):
    """Creates a chart and adds it to a new sheet."""
    source_sheet = workbook[sheet_name]
    chart_sheet = workbook.create_sheet(title=chart_title)

    if chart_type == 'bar':
        chart = openpyxl.chart.BarChart()
    elif chart_type == 'line':
        chart = openpyxl.chart.LineChart()
    elif chart_type == 'pie':
        chart = openpyxl.chart.PieChart()

    data_cols = []
    for col in y_columns:
        for i, column_cell in enumerate(source_sheet.iter_cols(min_row=1, max_row=1)):
            if column_cell[0].value == col:
                data_cols.append(i + 1)
                break

    cat_col = 0
    for i, column_cell in enumerate(source_sheet.iter_cols(min_row=1, max_row=1)):
        if column_cell[0].value == x_column:
            cat_col = i + 1
            break

    data = openpyxl.chart.Reference(source_sheet, min_col=data_cols[0], min_row=2, max_row=source_sheet.max_row, max_col=data_cols[-1])
    cats = openpyxl.chart.Reference(source_sheet, min_col=cat_col, min_row=2, max_row=source_sheet.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = title

    chart_sheet.add_chart(chart, "A1")

    return workbook

# --- Main Application ---

def main():
    parser = argparse.ArgumentParser(
        description='A versatile CLI tool for automating Excel workflows.',
        formatter_class=argparse.RawTextHelpFormatter
    )
    subparsers = parser.add_subparsers(dest='action', required=True, help='The action to perform')

    # --- Filter Action Parser ---
    parser_filter = subparsers.add_parser('filter', help='Filter rows based on a column value')
    parser_filter.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_filter.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_filter.add_argument('--column', required=True, help='Column to filter on')
    parser_filter.add_argument('--value', required=True, help='Value to filter for')

    # --- Summarize Action Parser ---
    parser_summarize = subparsers.add_parser('summarize', help='Summarize data by grouping and aggregating')
    parser_summarize.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_summarize.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_summarize.add_argument('--group-by', required=True, help='Column to group by')
    parser_summarize.add_argument('--agg-col', required=True, help='Column to aggregate')
    parser_summarize.add_argument('--agg-func', required=True, help='Aggregation function (e.g., mean, sum)')

    # --- Calculate Action Parser ---
    parser_calculate = subparsers.add_parser('calculate', help='Calculate a new column using an expression')
    parser_calculate.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_calculate.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_calculate.add_argument('--new-col', required=True, help='Name of the new column')
    parser_calculate.add_argument('--expr', required=True, help='Pandas-compatible expression (e.g., "Salary * 1.1")')

    # --- Merge Action Parser ---
    parser_merge = subparsers.add_parser('merge', help='Merge two Excel files')
    parser_merge.add_argument('--input1', required=True, help='First input Excel file (left)')
    parser_merge.add_argument('--input2', required=True, help='Second input Excel file (right)')
    parser_merge.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_merge.add_argument('--on', required=True, help='Column to merge on')
    parser_merge.add_argument('--how', default='inner', choices=['inner', 'outer', 'left', 'right'], help='Type of merge')

    # --- Sort Action Parser ---
    parser_sort = subparsers.add_parser('sort', help='Sort rows based on columns')
    parser_sort.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_sort.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_sort.add_argument('--by', required=True, nargs='+', help='Column(s) to sort by')
    parser_sort.add_argument('--order', default='asc', choices=['asc', 'desc'], help='Sort order')

    # --- Rename Columns Action Parser ---
    parser_rename = subparsers.add_parser('rename', help='Rename one or more columns')
    parser_rename.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_rename.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_rename.add_argument('--map', required=True, help='Mapping of old to new names (e.g., "OldName:NewName,Another:New")')

    # --- Drop Duplicates Action Parser ---
    parser_drop = subparsers.add_parser('drop_duplicates', help='Remove duplicate rows')
    parser_drop.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_drop.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_drop.add_argument('--subset', nargs='+', help='Column(s) to consider for identifying duplicates')

    # --- Duplicate Sheet Action Parser ---
    parser_duplicate = subparsers.add_parser('duplicate_sheet', help='Duplicate a sheet in an Excel file')
    parser_duplicate.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_duplicate.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_duplicate.add_argument('--source-sheet', required=True, help='Name of the sheet to duplicate')
    parser_duplicate.add_argument('--new-sheet-name', required=True, help='Name for the new duplicated sheet')

    # --- Update Cells Action Parser ---
    parser_update = subparsers.add_parser('update_cells', help='Update one or more cells in a sheet')
    parser_update.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_update.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_update.add_argument('--sheet-name', required=True, help='Name of the sheet to update')
    parser_update.add_argument('--updates', required=True, help='Cell updates in the format "A1:NewValue,B2:AnotherValue"')

    # --- Data Validation Action Parser ---
    parser_data_validation = subparsers.add_parser('data_validation', help='Perform data cleaning and validation operations')
    parser_data_validation.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_data_validation.add_argument('-o', '--output', required=True, help='Output Excel file')
    
    data_validation_subparsers = parser_data_validation.add_subparsers(dest='validation_action', required=True, help='Data validation operation')

    # Fill NA sub-action
    parser_fill_na = data_validation_subparsers.add_parser('fill_na', help='Fill missing values')
    parser_fill_na.add_argument('--value', required=True, help='Value to fill missing entries with')
    parser_fill_na.add_argument('--columns', nargs='+', help='Columns to fill NA values in (default: all columns)')

    # Convert Type sub-action
    parser_convert_type = data_validation_subparsers.add_parser('convert_type', help='Convert column data type')
    parser_convert_type.add_argument('--column', required=True, help='Column to convert')
    parser_convert_type.add_argument('--to-type', required=True, choices=['int', 'float', 'str', 'datetime'], help='Target data type')

    # --- Chart Action Parser ---
    parser_chart = subparsers.add_parser('chart', help='Create a chart from data')
    parser_chart.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser_chart.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser_chart.add_argument('--sheet-name', required=True, help='Sheet to draw chart from')
    parser_chart.add_argument('--chart-type', required=True, choices=['bar', 'line', 'pie'], help='Type of chart to create')
    parser_chart.add_argument('--x-column', required=True, help='Column for the X-axis (categories)')
    parser_chart.add_argument('--y-columns', required=True, nargs='+', help='Column(s) for the Y-axis (values)')
    parser_chart.add_argument('--title', default='Chart', help='Title of the chart')
    parser_chart.add_argument('--chart-title', default='Chart Sheet', help='Name of the new sheet for the chart')


    args = parser.parse_args()

    # --- Action Dispatch ---
    try:
        # Actions that modify the workbook directly with openpyxl
        if args.action in ['duplicate_sheet', 'update_cells', 'chart']:
            workbook = openpyxl.load_workbook(args.input)
            if args.action == 'duplicate_sheet':
                workbook = duplicate_sheet_data(workbook, args.source_sheet, args.new_sheet_name)
            elif args.action == 'update_cells':
                cell_updates = dict(item.split(':', 1) for item in args.updates.split(','))
                workbook = update_cells_data(workbook, args.sheet_name, cell_updates)
            elif args.action == 'chart':
                workbook = create_chart(workbook, args.sheet_name, args.chart_type, args.x_column, args.y_columns, args.title, args.chart_title)
            workbook.save(args.output)

        # Actions that process data with pandas
        else:
            if args.action == 'merge':
                df1 = pd.read_excel(args.input1)
                df2 = pd.read_excel(args.input2)
                result_df = merge_data(df1, df2, args.on, args.how)
            else: # For all other pandas-based actions
                df = pd.read_excel(args.input) # Read input for these actions
                if args.action == 'filter':
                    result_df = filter_data(df, args.column, args.value)
                elif args.action == 'summarize':
                    result_df = summarize_data(df, args.group_by, args.agg_col, args.agg_func)
                elif args.action == 'calculate':
                    result_df = calculate_column(df, args.new_col, args.expr)
                elif args.action == 'sort':
                    result_df = sort_data(df, args.by, ascending=(args.order == 'asc'))
                elif args.action == 'rename':
                    rename_map = dict(item.split(':') for item in args.map.split(','))
                    result_df = rename_columns_data(df, rename_map)
                elif args.action == 'drop_duplicates':
                    result_df = drop_duplicates_data(df, subset=args.subset)
                elif args.action == 'data_validation':
                    if args.validation_action == 'fill_na':
                        result_df = fill_missing_values(df, args.columns, args.value)
                    elif args.validation_action == 'convert_type':
                        result_df = convert_column_type(df, args.column, args.to_type)
                    else:
                        result_df = df # Should not happen due to required=True on subparsers
            
            result_df.to_excel(args.output, index=False)

        print(f"Action '{args.action}' completed successfully. Output saved to {args.output}")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == '__main__':
    main()
